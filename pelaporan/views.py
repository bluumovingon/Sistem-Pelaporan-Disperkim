from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, Http404, HttpResponseRedirect
from django.db.models import Q, Count
from django.utils import timezone
from django.urls import reverse
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from pelaporan.models import User, Kegiatan, Laporan, Dokumentasi, RiwayatStatus, Notifikasi
from pelaporan.forms import LaporanForm, KegiatanForm, UserForm
from pelaporan.decorators import role_required
from pelaporan.utils import compress_image, validate_file_signature
from django.db import transaction
import os

# --- Auth & Redirects ---

def home_redirect(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
    return redirect('login')


from django.core.cache import cache

def get_client_ip(request):
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0].strip()
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip

def login_view(request):
    if request.user.is_authenticated:
        return redirect('dashboard')
        
    # Rate Limiting Lockout berdasarkan IP
    ip = get_client_ip(request)
    lockout_key = f"lockout_{ip}"
    attempts_key = f"attempts_{ip}"
    
    is_locked = cache.get(lockout_key)
    if is_locked:
        messages.error(request, "Terlalu banyak kegagalan login. Akses Anda dikunci selama 5 menit.")
        return render(request, 'auth/login.html')
        
    if request.method == 'POST':
        u = request.POST.get('username')
        p = request.POST.get('password')
        
        user = authenticate(username=u, password=p)
        if user is not None:
            if user.status_aktif:
                cache.delete(attempts_key) # Reset percobaan jika sukses
                login(request, user)
                messages.success(request, f"Selamat datang kembali, {user.first_name}!")
                return redirect('dashboard')
            else:
                messages.error(request, "Akun Anda dinonaktifkan. Silakan hubungi Super Admin.")
        else:
            attempts = cache.get(attempts_key, 0) + 1
            cache.set(attempts_key, attempts, timeout=300) # Simpan selama 5 menit
            
            if attempts >= 5:
                cache.set(lockout_key, True, timeout=300) # Kunci IP selama 5 menit
                messages.error(request, "Terlalu banyak kegagalan login. Akses Anda dikunci selama 5 menit.")
            else:
                messages.error(request, f"Username atau password salah. Percobaan tersisa: {5 - attempts}")
            
    return render(request, 'auth/login.html')


@login_required
def logout_view(request):
    logout(request)
    messages.success(request, "Anda telah berhasil keluar dari sistem.")
    return redirect('login')


# --- Dashboard ---

@login_required
def dashboard_view(request):
    user = request.user
    
    # Prefetch untuk optimasi N+1 query pada riwayat laporan kegiatan
    from django.db.models import Prefetch
    latest_reports_prefetch = Prefetch(
        'laporan',
        queryset=Laporan.objects.order_by('-created_at'),
        to_attr='prefetched_laporans'
    )
    
    if user.role == 'pptk':
        # Dashboard PPTK
        laporans = Laporan.objects.filter(pptk=user)
        
        stats = {
            'total': laporans.count(),
            'draft': laporans.filter(status='draft').count(),
            'diajukan': laporans.filter(status='diajukan').count(),
            'terverifikasi': laporans.filter(status='terverifikasi').count(),
            'perlu_revisi': laporans.filter(status='perlu_revisi').count(),
        }
        
        recent_laporans = laporans.order_by('-updated_at')[:5]
        
        # Peta Spasial Kegiatan PPTK (Dioptimalkan dengan select_related & prefetch_related)
        kegiatans_query = Kegiatan.objects.filter(
            pptk=user, 
            latitude__isnull=False, 
            longitude__isnull=False
        ).select_related('pptk').prefetch_related(latest_reports_prefetch)
        
        kegiatans_list = []
        for k in kegiatans_query:
            latest_laporan = k.prefetched_laporans[0] if k.prefetched_laporans else None
            status = latest_laporan.status if latest_laporan else 'belum_ada'
            status_display = latest_laporan.get_status_display() if latest_laporan else 'Belum ada laporan'
            kegiatans_list.append({
                'id': k.id,
                'judul': k.judul_kegiatan,
                'tahun': k.tahun,
                'pptk_nama': f"{k.pptk.first_name} {k.pptk.last_name}",
                'latitude': k.latitude,
                'longitude': k.longitude,
                'status': status,
                'status_display': status_display,
            })
        
        context = {
            'stats': stats,
            'recent_laporans': recent_laporans,
            'kegiatans_list': kegiatans_list,
        }
        return render(request, 'dashboard/pptk.html', context)
        
    else:
        # Dashboard Admin / Pimpinan / Super Admin
        laporans = Laporan.objects.all()
        
        stats = {
            'total': laporans.count(),
            'diajukan': laporans.filter(status='diajukan').count(),
            'terverifikasi': laporans.filter(status='terverifikasi').count(),
            'perlu_revisi': laporans.filter(status='perlu_revisi').count(),
        }
        
        # Laporan yang butuh verifikasi (Optimasi select_related)
        pending_reports = laporans.filter(status='diajukan').select_related('pptk', 'kegiatan').order_by('-updated_at')
        
        # Kegiatan baru yang diajukan oleh PPTK (ACC kegiatan oleh Admin)
        pending_kegiatans = Kegiatan.objects.filter(status='diajukan').select_related('pptk').order_by('-tahun')
        
        # Chart Data
        # 1. Reports by status
        status_chart_data = list(laporans.values('status').annotate(count=Count('id')))
        # 2. Reports by Month
        month_chart_data = list(laporans.values('bulan_laporan').annotate(count=Count('id')))
        # 3. Reports by PPTK
        pptk_chart_data = list(laporans.values('pptk__first_name', 'pptk__last_name').annotate(count=Count('id')))
        
        # Peta Spasial Seluruh Kegiatan (Optimasi select_related & prefetch_related)
        kegiatans_query = Kegiatan.objects.filter(
            latitude__isnull=False, 
            longitude__isnull=False
        ).select_related('pptk').prefetch_related(latest_reports_prefetch)
        
        kegiatans_list = []
        for k in kegiatans_query:
            latest_laporan = k.prefetched_laporans[0] if k.prefetched_laporans else None
            status = latest_laporan.status if latest_laporan else 'belum_ada'
            status_display = latest_laporan.get_status_display() if latest_laporan else 'Belum ada laporan'
            kegiatans_list.append({
                'id': k.id,
                'judul': k.judul_kegiatan,
                'tahun': k.tahun,
                'pptk_nama': f"{k.pptk.first_name} {k.pptk.last_name}",
                'latitude': k.latitude,
                'longitude': k.longitude,
                'status': status,
                'status_display': status_display,
            })
        
        context = {
            'stats': stats,
            'pending_reports': pending_reports[:10],
            'pending_kegiatans': pending_kegiatans[:10],
            'status_chart_data': status_chart_data,
            'month_chart_data': month_chart_data,
            'pptk_chart_data': pptk_chart_data,
            'kegiatans_list': kegiatans_list,
        }
        return render(request, 'dashboard/admin.html', context)


# --- Laporan CRUD ---

@login_required
def laporan_list_view(request):
    user = request.user
    
    # Base query based on roles (Optimasi N+1 select_related)
    if user.role == 'pptk':
        laporans = Laporan.objects.filter(pptk=user).select_related('pptk', 'kegiatan')
    else:
        laporans = Laporan.objects.all().select_related('pptk', 'kegiatan')
        
    # Apply Filters
    q_status = request.GET.get('status')
    q_bulan = request.GET.get('bulan')
    q_kegiatan = request.GET.get('kegiatan')
    q_pptk = request.GET.get('pptk')
    q_search = request.GET.get('search')
    
    if q_status:
        laporans = laporans.filter(status=q_status)
    if q_bulan:
        laporans = laporans.filter(bulan_laporan=q_bulan)
    if q_kegiatan:
        laporans = laporans.filter(kegiatan_id=q_kegiatan)
    if q_pptk and user.role != 'pptk':
        laporans = laporans.filter(pptk_id=q_pptk)
    if q_search:
        laporans = laporans.filter(
            Q(kegiatan__judul_kegiatan__icontains=q_search) | 
            Q(tahapan_pelaksanaan__icontains=q_search) | 
            Q(kendala__icontains=q_search)
        )
        
    # Urutkan agar konsisten dalam pagination
    laporans = laporans.order_by('-updated_at')
    
    # Pagination (10 data per halaman)
    from django.core.paginator import Paginator
    paginator = Paginator(laporans, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get metadata for filter inputs
    if user.role == 'pptk':
        all_kegiatans = Kegiatan.objects.filter(pptk=user, status='disetujui')
    else:
        all_kegiatans = Kegiatan.objects.all()
    all_pptk = User.objects.filter(role='pptk')
    all_bulans = [b[0] for b in Laporan.BULAN_CHOICES]
    all_statuses = Laporan.STATUS_CHOICES
    
    context = {
        'laporans': page_obj,
        'page_obj': page_obj,
        'all_kegiatans': all_kegiatans,
        'all_pptk': all_pptk,
        'all_bulans': all_bulans,
        'all_statuses': all_statuses,
        'filters': {
            'status': q_status or '',
            'bulan': q_bulan or '',
            'kegiatan': q_kegiatan or '',
            'pptk': q_pptk or '',
            'search': q_search or '',
        }
    }
    return render(request, 'laporan/list.html', context)
@login_required
@role_required('pptk')
@transaction.atomic
def laporan_buat_view(request):
    if request.method == 'POST':
        form = LaporanForm(request.POST)
        form.fields['kegiatan'].queryset = Kegiatan.objects.filter(pptk=request.user, status='disetujui')
        # Validate files first
        files = request.FILES.getlist('files')
        allowed_extensions = ['.jpg', '.jpeg', '.png', '.pdf']
        max_size = 5 * 1024 * 1024 # 5MB
        
        file_errors = []
        for f in files:
            ext = os.path.splitext(f.name)[1].lower() if '.' in f.name else ''
            if ext not in allowed_extensions:
                file_errors.append(f"Berkas '{f.name}' memiliki format '{ext}' yang tidak diizinkan. (Hanya JPG, JPEG, PNG, PDF)")
            elif not validate_file_signature(f):
                file_errors.append(f"Berkas '{f.name}' tidak valid atau isi berkas tidak cocok dengan ekstensinya.")
            if f.size > max_size:
                file_errors.append(f"Berkas '{f.name}' melebihi batas ukuran 5MB.")
                
        if file_errors:
            for err in file_errors:
                messages.error(request, err)
            return render(request, 'laporan/form.html', {'form': form, 'is_new': True})

        if form.is_valid():
            laporan = form.save(commit=False)
            laporan.pptk = request.user
            
            # Check action (draft or submit)
            action = request.POST.get('action')
            if action == 'diajukan':
                laporan.status = 'diajukan'
            else:
                laporan.status = 'draft'
                
            laporan.save()
            
            # Save historical log
            RiwayatStatus.objects.create(
                laporan=laporan,
                status_lama='draft', # dummy initial status
                status_baru=laporan.status,
                diubah_oleh=request.user,
                catatan='Laporan dibuat pertama kali.'
            )
            
            # Save files
            file_descriptions = request.POST.getlist('file_descriptions')
            for i, f in enumerate(files):
                desc = file_descriptions[i] if i < len(file_descriptions) else ''
                compressed_f = compress_image(f)
                Dokumentasi.objects.create(
                    laporan=laporan,
                    file=compressed_f,
                    keterangan=desc
                )
                
            # If submitted directly, notify admins
            if action == 'diajukan':
                admins = User.objects.filter(role__in=['admin', 'super_admin'])
                for admin in admins:
                    Notifikasi.objects.create(
                        user=admin,
                        judul="Laporan Baru Diajukan",
                        pesan=f"Laporan kegiatan '{laporan.kegiatan.judul_kegiatan}' bulan {laporan.bulan_laporan} telah diajukan oleh {request.user.first_name} {request.user.last_name}.",
                        laporan=laporan
                    )
                    
            messages.success(request, f"Laporan berhasil disimpan sebagai {laporan.get_status_display()}.")
            return redirect('laporan_list')
    else:
        form = LaporanForm()
        form.fields['kegiatan'].queryset = Kegiatan.objects.filter(pptk=request.user, status='disetujui')
        
    return render(request, 'laporan/form.html', {'form': form, 'is_new': True})


@login_required
def laporan_detail_view(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    
    # Authorization check
    if request.user.role == 'pptk' and laporan.pptk != request.user:
        messages.error(request, "Anda tidak diperbolehkan melihat laporan ini.")
        return redirect('dashboard')
        
    riwayats = laporan.riwayat_status.all().order_by('-created_at')
    dokumentasis = laporan.dokumentasi.all()
    
    context = {
        'laporan': laporan,
        'riwayats': riwayats,
        'dokumentasis': dokumentasis,
    }
    return render(request, 'laporan/detail.html', context)
@login_required
@role_required('pptk')
@transaction.atomic
def laporan_edit_view(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    
    # Ownership & status constraint checks
    if laporan.pptk != request.user:
        messages.error(request, "Anda hanya bisa mengedit laporan milik sendiri.")
        return redirect('dashboard')
        
    if not (laporan.status == 'draft' or laporan.status == 'perlu_revisi'):
        messages.error(request, "Laporan yang sudah diajukan atau diverifikasi tidak bisa diedit.")
        return redirect('laporan_detail', pk=laporan.pk)
        
    if request.method == 'POST':
        form = LaporanForm(request.POST, instance=laporan)
        form.fields['kegiatan'].queryset = Kegiatan.objects.filter(pptk=request.user, status='disetujui')
        dokumentasis = laporan.dokumentasi.all()
        
        # Validate files first
        files = request.FILES.getlist('files')
        allowed_extensions = ['.jpg', '.jpeg', '.png', '.pdf']
        max_size = 5 * 1024 * 1024 # 5MB
        
        file_errors = []
        for f in files:
            ext = os.path.splitext(f.name)[1].lower() if '.' in f.name else ''
            if ext not in allowed_extensions:
                file_errors.append(f"Berkas '{f.name}' memiliki format '{ext}' yang tidak diizinkan. (Hanya JPG, JPEG, PNG, PDF)")
            elif not validate_file_signature(f):
                file_errors.append(f"Berkas '{f.name}' tidak valid atau isi berkas tidak cocok dengan ekstensinya.")
            if f.size > max_size:
                file_errors.append(f"Berkas '{f.name}' melebihi batas ukuran 5MB.")
                
        if file_errors:
            for err in file_errors:
                messages.error(request, err)
            return render(request, 'laporan/form.html', {
                'form': form,
                'laporan': laporan,
                'dokumentasis': dokumentasis,
                'is_new': False
            })

        if form.is_valid():
            original_status = laporan.status
            laporan = form.save(commit=False)
            laporan.pptk = request.user
            
            # Check action
            action = request.POST.get('action')
            if action == 'diajukan':
                laporan.status = 'diajukan'
                
            laporan.save()
            
            # Save status log if status changed
            if original_status != laporan.status:
                RiwayatStatus.objects.create(
                    laporan=laporan,
                    status_lama=original_status,
                    status_baru=laporan.status,
                    diubah_oleh=request.user,
                    catatan='Laporan diajukan kembali setelah diedit/revisi.'
                )
                
                # Notify admins
                admins = User.objects.filter(role__in=['admin', 'super_admin'])
                for admin in admins:
                    Notifikasi.objects.create(
                        user=admin,
                        judul="Laporan Pengajuan Ulang",
                        pesan=f"Laporan kegiatan '{laporan.kegiatan.judul_kegiatan}' bulan {laporan.bulan_laporan} diajukan kembali oleh {request.user.first_name} {request.user.last_name}.",
                        laporan=laporan
                    )
            
            # Save files
            file_descriptions = request.POST.getlist('file_descriptions')
            for i, f in enumerate(files):
                desc = file_descriptions[i] if i < len(file_descriptions) else ''
                compressed_f = compress_image(f)
                Dokumentasi.objects.create(
                    laporan=laporan,
                    file=compressed_f,
                    keterangan=desc
                )
                
            messages.success(request, f"Laporan berhasil diperbarui.")
            return redirect('laporan_detail', pk=laporan.pk)
    else:
        form = LaporanForm(instance=laporan)
        form.fields['kegiatan'].queryset = Kegiatan.objects.filter(pptk=request.user, status='disetujui')
        
    dokumentasis = laporan.dokumentasi.all()
    context = {
        'form': form,
        'laporan': laporan,
        'dokumentasis': dokumentasis,
        'is_new': False
    }
    return render(request, 'laporan/form.html', context)


@login_required
@role_required('admin', 'super_admin')
@transaction.atomic
def laporan_verifikasi_view(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        catatan = request.POST.get('catatan_revisi', '')
        
        status_lama = laporan.status
        
        if action == 'approve':
            laporan.status = 'terverifikasi'
            laporan.catatan_revisi = ''
            laporan.save()
            
            # Histori
            RiwayatStatus.objects.create(
                laporan=laporan,
                status_lama=status_lama,
                status_baru='terverifikasi',
                diubah_oleh=request.user,
                catatan='Laporan disetujui dan diverifikasi.'
            )
            
            # Notify pptk
            Notifikasi.objects.create(
                user=laporan.pptk,
                judul="Laporan Disetujui",
                pesan=f"Laporan kegiatan '{laporan.kegiatan.judul_kegiatan}' untuk bulan {laporan.bulan_laporan} telah disetujui oleh Admin.",
                laporan=laporan
            )
            
            messages.success(request, "Laporan berhasil diverifikasi/disetujui.")
            
        elif action == 'reject':
            if not catatan:
                messages.error(request, "Catatan alasan revisi harus diisi.")
                return redirect('laporan_detail', pk=laporan.pk)
                
            laporan.status = 'perlu_revisi'
            laporan.catatan_revisi = catatan
            laporan.save()
            
            # Histori
            RiwayatStatus.objects.create(
                laporan=laporan,
                status_lama=status_lama,
                status_baru='perlu_revisi',
                diubah_oleh=request.user,
                catatan=f"Laporan ditolak / perlu revisi. Catatan: {catatan}"
            )
            
            # Notify pptk
            Notifikasi.objects.create(
                user=laporan.pptk,
                judul="Laporan Butuh Revisi",
                pesan=f"Laporan kegiatan '{laporan.kegiatan.judul_kegiatan}' bulan {laporan.bulan_laporan} ditolak/butuh revisi. Catatan: {catatan}",
                laporan=laporan
            )
            
            messages.warning(request, "Laporan ditolak dan dikembalikan ke PPTK untuk direvisi.")
            
    return redirect('laporan_detail', pk=laporan.pk)


@login_required
def laporan_hapus_view(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    
    # Laporan terverifikasi tidak boleh dihapus oleh siapa pun (termasuk admin) demi integritas audit
    if laporan.status == 'terverifikasi':
        messages.error(request, "Laporan yang sudah Terverifikasi tidak dapat dihapus untuk menjaga keaslian jejak audit.")
        return redirect('laporan_detail', pk=laporan.pk)
        
    # Auth constraint
    if request.user.role == 'pimpinan':
        messages.error(request, "Anda tidak diperbolehkan menghapus laporan.")
        return redirect('dashboard')

    if request.user.role == 'pptk':
        if laporan.pptk != request.user:
            messages.error(request, "Anda tidak diizinkan menghapus laporan ini.")
            return redirect('dashboard')
        if laporan.status not in ['draft', 'perlu_revisi']:
            messages.error(request, "Anda hanya bisa menghapus laporan berstatus Draft atau Perlu Revisi.")
            return redirect('laporan_detail', pk=laporan.pk)
        
    # Delete file fields from disk
    docs = laporan.dokumentasi.all()
    for doc in docs:
        if doc.file:
            try:
                doc.file.delete(save=False)
            except Exception:
                pass
                
    laporan.delete()
    messages.success(request, "Laporan berhasil dihapus.")
    return redirect('laporan_list')


@login_required
def dokumentasi_hapus_view(request, pk):
    doc = get_object_or_404(Dokumentasi, pk=pk)
    laporan = doc.laporan
    
    # Auth constraint
    if request.user.role == 'pimpinan':
        messages.error(request, "Anda tidak diperbolehkan menghapus dokumentasi.")
        return redirect('dashboard')
        
    if request.user.role == 'pptk' and laporan.pptk != request.user:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('dashboard')))
        
    if request.user.role == 'pptk' and laporan.status not in ['draft', 'perlu_revisi']:
        return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('dashboard')))
        
    # Delete from disk
    if doc.file:
        try:
            doc.file.delete(save=False)
        except Exception:
            pass
            
    doc.delete()
    messages.success(request, "File dokumentasi berhasil dihapus.")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('laporan_edit', args=[laporan.pk])))


# --- Rekap & Exports ---

@login_required
@role_required('admin', 'pimpinan', 'super_admin')
def rekap_view(request):
    # Apply Filters
    q_bulan = request.GET.get('bulan')
    q_kegiatan = request.GET.get('kegiatan')
    q_pptk = request.GET.get('pptk')
    
    # We only show terverifikasi reports in the recap sheet
    laporans = Laporan.objects.filter(status='terverifikasi')
    
    if q_bulan:
        laporans = laporans.filter(bulan_laporan=q_bulan)
    if q_kegiatan:
        laporans = laporans.filter(kegiatan_id=q_kegiatan)
    if q_pptk:
        laporans = laporans.filter(pptk_id=q_pptk)
        
    all_kegiatans = Kegiatan.objects.all()
    all_pptk = User.objects.filter(role='pptk')
    all_bulans = [b[0] for b in Laporan.BULAN_CHOICES]
    
    context = {
        'laporans': laporans,
        'all_kegiatans': all_kegiatans,
        'all_pptk': all_pptk,
        'all_bulans': all_bulans,
        'filters': {
            'bulan': q_bulan or '',
            'kegiatan': q_kegiatan or '',
            'pptk': q_pptk or '',
        }
    }
    return render(request, 'rekap/rekap.html', context)


@login_required
@role_required('admin', 'pimpinan', 'super_admin')
def rekap_excel_view(request):
    q_bulan = request.GET.get('bulan')
    q_kegiatan = request.GET.get('kegiatan')
    q_pptk = request.GET.get('pptk')
    
    # Filter reports
    laporans = Laporan.objects.filter(status='terverifikasi')
    if q_bulan:
        laporans = laporans.filter(bulan_laporan=q_bulan)
    if q_kegiatan:
        laporans = laporans.filter(kegiatan_id=q_kegiatan)
    if q_pptk:
        laporans = laporans.filter(pptk_id=q_pptk)
        
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Rekap Laporan"
    
    # Grid lines visible
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.merge_cells("A1:H1")
    ws["A1"] = "REKAPITULASI LAPORAN KEGIATAN PENGAWASAN LAPANGAN"
    ws["A1"].font = Font(name="Calibri", size=14, bold=True, color="000000")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A2:H2")
    ws["A2"] = "DINAS PERUMAHAN DAN PERMUKIMAN (DISPERKIM)"
    ws["A2"].font = Font(name="Calibri", size=12, bold=True)
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells("A3:H3")
    filter_desc = "Semua Laporan Terverifikasi"
    if q_bulan:
        filter_desc += f" | Bulan: {q_bulan}"
    ws["A3"] = filter_desc
    ws["A3"].font = Font(name="Calibri", size=10, italic=True)
    ws["A3"].alignment = Alignment(horizontal="center", vertical="center")
    
    # Row Heights
    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 20
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[5].height = 26
    
    # Headers
    headers = [
        "No", 
        "Nama PPTK", 
        "Jabatan PPTK", 
        "Judul Kegiatan", 
        "Bulan Laporan", 
        "Tahapan Pelaksanaan", 
        "Kendala", 
        "Status"
    ]
    
    # Styling variables
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Write headers (row 5)
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=5, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
        
    # Write data rows
    current_row = 6
    for idx, lap in enumerate(laporans, 1):
        ws.row_dimensions[current_row].height = 35  # tall data rows for text wrapping
        
        c1 = ws.cell(row=current_row, column=1, value=idx)
        c1.alignment = center_align
        
        c2 = ws.cell(row=current_row, column=2, value=f"{lap.pptk.first_name} {lap.pptk.last_name}")
        c2.alignment = left_align
        
        c3 = ws.cell(row=current_row, column=3, value=lap.pptk.jabatan)
        c3.alignment = left_align
        
        c4 = ws.cell(row=current_row, column=4, value=lap.kegiatan.judul_kegiatan)
        c4.alignment = left_align
        
        c5 = ws.cell(row=current_row, column=5, value=lap.bulan_laporan)
        c5.alignment = center_align
        
        c6 = ws.cell(row=current_row, column=6, value=lap.tahapan_pelaksanaan)
        c6.alignment = left_align
        
        c7 = ws.cell(row=current_row, column=7, value=lap.kendala or "-")
        c7.alignment = left_align
        
        c8 = ws.cell(row=current_row, column=8, value=lap.get_status_display())
        c8.alignment = center_align
        
        # Apply borders and fonts to data cells
        for col_idx in range(1, 9):
            cell = ws.cell(row=current_row, column=col_idx)
            cell.font = Font(name="Calibri", size=10)
            cell.border = thin_border
            
        current_row += 1
        
    # Adjust column widths
    column_widths = {
        'A': 5,   # No
        'B': 25,  # Nama PPTK
        'C': 25,  # Jabatan PPTK
        'D': 35,  # Judul Kegiatan
        'E': 15,  # Bulan
        'F': 45,  # Tahapan
        'G': 30,  # Kendala
        'H': 15,  # Status
    }
    
    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width
        
    # Build response
    filename = f"rekap_sipawas_{timezone.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


@login_required
def rekap_pdf_view(request, pk):
    laporan = get_object_or_404(Laporan, pk=pk)
    
    # Auth constraint
    if request.user.role == 'pptk' and laporan.pptk != request.user:
        messages.error(request, "Anda tidak diperbolehkan mengakses cetak laporan ini.")
        return redirect('dashboard')
        
    dokumentasis = laporan.dokumentasi.all()
    riwayats = laporan.riwayat_status.all().order_by('created_at')
    
    context = {
        'laporan': laporan,
        'dokumentasis': dokumentasis,
        'riwayats': riwayats,
        'current_date': timezone.now()
    }
    return render(request, 'rekap/pdf_report.html', context)


# --- Master Kegiatan ---

@login_required
@role_required('admin', 'super_admin', 'pptk')
def kegiatan_list_view(request):
    user = request.user
    if user.role == 'pptk':
        kegiatans = Kegiatan.objects.filter(pptk=user).select_related('pptk').order_by('-id')
    else:
        kegiatans = Kegiatan.objects.all().select_related('pptk').order_by('-id')
        
    # Pagination (10 data per halaman)
    from django.core.paginator import Paginator
    paginator = Paginator(kegiatans, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'master/kegiatan_list.html', {
        'kegiatans': page_obj,
        'page_obj': page_obj
    })


@login_required
@role_required('admin', 'super_admin', 'pptk')
@transaction.atomic
def kegiatan_tambah_view(request):
    if request.method == 'POST':
        form = KegiatanForm(request.POST)
        if request.user.role == 'pptk':
            form.fields['pptk'].required = False
        if form.is_valid():
            kegiatan = form.save(commit=False)
            if request.user.role == 'pptk':
                kegiatan.pptk = request.user
                kegiatan.status = 'diajukan' # Diajukan untuk persetujuan admin
                success_msg = "Usulan Kegiatan berhasil ditambahkan dan menunggu persetujuan Admin."
            else:
                kegiatan.status = 'disetujui'
                success_msg = "Data Kegiatan berhasil ditambahkan."
            kegiatan.save()
            messages.success(request, success_msg)
            return redirect('kegiatan_list')
    else:
        form = KegiatanForm()
        if request.user.role == 'pptk':
            form.fields['pptk'].required = False
    return render(request, 'master/kegiatan_form.html', {'form': form, 'is_new': True})


@login_required
@role_required('admin', 'super_admin', 'pptk')
@transaction.atomic
def kegiatan_edit_view(request, pk):
    kegiatan = get_object_or_404(Kegiatan, pk=pk)
    if request.user.role == 'pptk' and kegiatan.pptk != request.user:
        messages.error(request, "Anda hanya dapat mengubah kegiatan milik sendiri.")
        return redirect('kegiatan_list')
        
    if request.method == 'POST':
        form = KegiatanForm(request.POST, instance=kegiatan)
        if request.user.role == 'pptk':
            form.fields['pptk'].required = False
        if form.is_valid():
            kegiatan = form.save(commit=False)
            if request.user.role == 'pptk':
                kegiatan.pptk = request.user
                kegiatan.status = 'diajukan' # Diajukan kembali untuk persetujuan admin setelah diedit oleh pptk
                success_msg = "Perubahan Kegiatan berhasil diajukan dan menunggu persetujuan Admin."
            else:
                kegiatan.status = 'disetujui'
                success_msg = "Data Kegiatan berhasil diperbarui."
            kegiatan.save()
            messages.success(request, success_msg)
            return redirect('kegiatan_list')
    else:
        form = KegiatanForm(instance=kegiatan)
        if request.user.role == 'pptk':
            form.fields['pptk'].required = False
    return render(request, 'master/kegiatan_form.html', {'form': form, 'is_new': False, 'kegiatan': kegiatan})


@login_required
@role_required('admin', 'super_admin', 'pptk')
def kegiatan_hapus_view(request, pk):
    kegiatan = get_object_or_404(Kegiatan, pk=pk)
    if request.user.role == 'pptk' and kegiatan.pptk != request.user:
        messages.error(request, "Anda hanya dapat menghapus kegiatan milik sendiri.")
        return redirect('kegiatan_list')
        
    if kegiatan.laporan.exists():
        messages.error(request, "Tidak dapat menghapus Kegiatan karena masih terhubung dengan laporan aktif.")
    else:
        kegiatan.delete()
        messages.success(request, "Master data Kegiatan berhasil dihapus.")
    return redirect('kegiatan_list')


# --- Master Pengguna (Super Admin) ---

@login_required
@role_required('super_admin')
def pengguna_list_view(request):
    penggunas = User.objects.all().order_by('-id')
    
    # Pagination (10 data per halaman)
    from django.core.paginator import Paginator
    paginator = Paginator(penggunas, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'master/pengguna_list.html', {
        'penggunas': page_obj,
        'page_obj': page_obj
    })


@login_required
@role_required('super_admin')
def pengguna_tambah_view(request):
    if request.method == 'POST':
        form = UserForm(request.POST)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = user.role in ['admin', 'super_admin']
            user.is_superuser = user.role == 'super_admin'
            user.save()
            messages.success(request, f"Pengguna '{user.username}' berhasil ditambahkan.")
            return redirect('pengguna_list')
    else:
        form = UserForm()
    return render(request, 'master/pengguna_form.html', {'form': form, 'is_new': True})


@login_required
@role_required('super_admin')
def pengguna_edit_view(request, pk):
    user = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        form = UserForm(request.POST, instance=user)
        if form.is_valid():
            user = form.save(commit=False)
            user.is_staff = user.role in ['admin', 'super_admin']
            user.is_superuser = user.role == 'super_admin'
            user.save()
            messages.success(request, f"Pengguna '{user.username}' berhasil diperbarui.")
            return redirect('pengguna_list')
    else:
        form = UserForm(instance=user)
    return render(request, 'master/pengguna_form.html', {'form': form, 'is_new': False, 'target_user': user})


@login_required
@role_required('super_admin')
def pengguna_hapus_view(request, pk):
    user = get_object_or_404(User, pk=pk)
    if user == request.user:
        messages.error(request, "Anda tidak dapat menghapus akun Anda sendiri yang sedang aktif.")
    else:
        username = user.username
        user.delete()
        messages.success(request, f"Pengguna '{username}' berhasil dihapus.")
    return redirect('pengguna_list')


# --- Notifikasi ---

@login_required
def notifikasi_list_view(request):
    notifs = Notifikasi.objects.filter(user=request.user)
    return render(request, 'notifikasi/list.html', {'notifs': notifs})


@login_required
def notifikasi_baca_view(request, pk):
    notif = get_object_or_404(Notifikasi, pk=pk, user=request.user)
    notif.is_read = True
    notif.save()
    if notif.laporan:
        return redirect('laporan_detail', pk=notif.laporan.pk)
    return redirect('notifikasi_list')


@login_required
def notifikasi_baca_semua_view(request):
    Notifikasi.objects.filter(user=request.user, is_read=False).update(is_read=True)
    messages.success(request, "Semua notifikasi telah ditandai sebagai dibaca.")
    return HttpResponseRedirect(request.META.get('HTTP_REFERER', reverse('dashboard')))


# --- Proteksi Berkas Media & Verifikasi Kegiatan ---

from django.http import FileResponse, Http404
import mimetypes

@login_required
def serve_dokumentasi_view(request, filename):
    filename = os.path.basename(filename)
    doc = get_object_or_404(Dokumentasi, file='dokumentasi/' + filename)
    laporan = doc.laporan
    
    # Cek otorisasi
    if request.user.role == 'pptk' and laporan.pptk != request.user:
        raise Http404("Anda tidak diizinkan mengakses berkas ini.")
        
    file_path = doc.file.path
    if not os.path.exists(file_path):
        raise Http404("Berkas tidak ditemukan.")
        
    content_type, _ = mimetypes.guess_type(file_path)
    if not content_type:
        content_type = 'application/octet-stream'
        
    return FileResponse(open(file_path, 'rb'), content_type=content_type)


@login_required
@role_required('admin', 'super_admin')
@transaction.atomic
def kegiatan_verifikasi_view(request, pk):
    kegiatan = get_object_or_404(Kegiatan, pk=pk)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        catatan = request.POST.get('catatan_admin', '')
        
        if action == 'approve':
            kegiatan.status = 'disetujui'
            kegiatan.catatan_admin = ''
            kegiatan.save()
            
            Notifikasi.objects.create(
                user=kegiatan.pptk,
                judul="Kegiatan Disetujui",
                pesan=f"Usulan Kegiatan/Proyek '{kegiatan.judul_kegiatan}' Anda telah disetujui (ACC) oleh Admin.",
                laporan=None
            )
            messages.success(request, "Kegiatan berhasil disetujui.")
            
        elif action == 'reject':
            if not catatan:
                messages.error(request, "Catatan/alasan penolakan harus diisi.")
                return redirect('kegiatan_list')
                
            kegiatan.status = 'ditolak'
            kegiatan.catatan_admin = catatan
            kegiatan.save()
            
            Notifikasi.objects.create(
                user=kegiatan.pptk,
                judul="Kegiatan Ditolak",
                pesan=f"Usulan Kegiatan/Proyek '{kegiatan.judul_kegiatan}' Anda ditolak oleh Admin. Catatan: {catatan}",
                laporan=None
            )
            messages.warning(request, "Usulan Kegiatan telah ditolak.")
            
    return redirect('kegiatan_list')

